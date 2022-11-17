import pickle
import os
import base64

import shutil
import time
import datetime
from email import encoders
from email.mime.base import MIMEBase
from pathlib import Path
import googleapiclient.discovery
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import numpy as np
import openpyxl
import winsound


def flipexcel(test):
    if test == 1:
        """ Defining the path of the rtgfnxflip file which will read and write in."""
        xlsx_file4 = Path('fnxlist', 'rtgfnxflip.xlsx')
        wb_obj4 = openpyxl.load_workbook(xlsx_file4, data_only=True)
        rtgfnxflip = wb_obj4.active
        row_countfnxflip = rtgfnxflip.max_row
        column_countfnxflip = rtgfnxflip.max_column
        maxfnxflip = 0
        temp11flip = 0

        """ Defining the path of the rtgfnx file which will read and write in."""
        xlsx_file1 = Path('fnxlist', 'rtgfnx.xlsx')
        wb_obj1 = openpyxl.load_workbook(xlsx_file1, data_only=True)
        rtgfnx = wb_obj1.active
        row_countfnx = rtgfnx.max_row
        column_countfnx = rtgfnx.max_column
        maxfnx = 0
        temp11 = 0

        """ Just to be sure that the maxfnx ID is updated after every change"""
        for i in range(9, row_countfnx):
            temp16 = rtgfnx["A" + str(i)].value
            if temp16 is not None:
                # print(temp16, type(temp16))
                if maxfnx < temp16:
                    maxfnx = int(temp16)
        c = 0
        for i in range(1, row_countfnx):
            # print(i)
            rtgfnxflip['A' + str(i+1)] = rtgfnx['A' + str(i+8)].value
            rtgfnxflip['B' + str(i+1)] = rtgfnx['B' + str(i+8)].value
            rtgfnxflip['H' + str(i+1)] = rtgfnx['J' + str(i+8)].value
        wb_obj4.save(xlsx_file4)
    else:
        print("WRONG USAGE OF DEBUGFILE.PY, test != 1")
        return 0
    return 0


print("BACKGROUND FILE, DOES NOT INTERACT WITH THE USER")
print("PLEASE OPEN XLSXREADER.PY OR CONTACT @DAVI ALVES DO NASCIMENTO")
# flipexcel(1)
# gmail()
duration = 1000  # milliseconds

freq = 440  # Hz
winsound.Beep(freq, duration)
print(str(datetime.datetime.now()))

print("flipend")



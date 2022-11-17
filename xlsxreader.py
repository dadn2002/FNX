import os
import shutil
import time
import winsound
from datetime import datetime
from pathlib import Path
# from debugfile import flipexcel

import numpy as np
import openpyxl


def languagereduce(str1, str2):
    str3 = str1.lower()
    str4 = str2.lower()
    for ij in range(0, len(str3)):
        if ord(str3[ij]) == 135:
            str3[ij] = 'c'
        elif ord(str3[ij]) == (131 or 132 or 133 or 134 or 160):
            str3[ij] = 'a'
        elif ord(str3[ij]) == (130 or 136 or 137):
            str3[ij] = 'e'
        elif ord(str3[ij]) == (139 or 140 or 141 or 161):
            str3[ij] = 'i'
        elif ord(str3[ij]) == (147 or 148 or 149 or 162):
            str3[ij] = 'o'
        elif ord(str3[ij]) == (129 or 150 or 151 or 163):
            str3[ij] = 'u'
    for ij in range(0, len(str4)):
        if ord(str4[ij]) == 135:
            str4[ij] = 'c'
        elif ord(str4[ij]) == (131 or 132 or 133 or 134 or 160):
            str4[ij] = 'a'
        elif ord(str4[ij]) == (130 or 136 or 137):
            str4[ij] = 'e'
        elif ord(str4[ij]) == (139 or 140 or 141 or 161):
            str4[ij] = 'i'
        elif ord(str4[ij]) == (147 or 148 or 149 or 162):
            str4[ij] = 'o'
        elif ord(str4[ij]) == (129 or 150 or 151 or 163):
            str4[ij] = 'u'
    str5 = str3.split(' ')
    cn = 0
    for ij in range(len(str5)):
        if (str5[ij] in str4) and len(str5[ij]) > 2:
            cn = cn + 1
    # print("debug", str5, str4, cn)
    return cn


def countingwords(str1, str2):
    str3 = str1.split(' ')
    cm = 0
    for iz in range(0, len(str3)):
        if str3[iz] in str2:
            if len(str3[iz]) > 2:
                cm = cm + 1
    if cm > 2:
        return True
    return False


def var(rtg1, rtg2, result, t):  # Function rating deviation
    """ Return the rating variation given player1 rtg, player2 rtg, result and time control,
     the formula can be found in https://en.wikipedia.org/wiki/Elo_rating_system"""
    k = 0  # Setting K value for classical/rapid/blitz
    if t == 0:
        k = 25
    elif t == 1:
        k = 15
    elif t == 2:
        k = 10
    else:
        return 0

    if rtg1 - rtg2 > 400:  # Capping rating dif at 400
        rtg1 = rtg2 + 400
    elif rtg2 - rtg1 > 400:
        rtg2 = rtg1 + 400

    Q1 = pow(10, (rtg1 / 400))
    Q2 = pow(10, (rtg2 / 400))
    E1 = Q1 / (Q1 + Q2)
    # E2 = Q2/(Q1+Q2)
    result = str(result)
    if ord(result) == 189:  # Actually ½ isn't a number for some reason
        result = 0.5
    # print(result)
    if result != '1' and result != 0.5 and result != '0':
        result = 0
    result = float(result)
    if rtg1 == 0:
        return 0
    if rtg2 == 0:
        return 0
    return round(float(k * (float(result) - E1)), 2)  # Rating deviation


def findplayer(name, size, where, hint=None):
    # print("Debug2:", name, size, where, hint)
    """ Return the FNX ID, Name and FNX Rtg of player in a vector len 3."""
    # xlsx_file1a = Path('fnxlist', 'rtgfnx.xlsx')
    # wb_obj1a = openpyxl.load_workbook(xlsx_file1a, data_only=True)
    # rtgfnx1a = wb_obj1a.active
    name = name.lower()
    for ij in range(0, len(name)):
        if ord(name[ij]) == 135:
            name[ij] = 'c'
        elif ord(name[ij]) == (131 or 132 or 133 or 134 or 160):
            name[ij] = 'a'
        elif ord(name[ij]) == (130 or 136 or 137):
            name[ij] = 'e'
        elif ord(name[ij]) == (139 or 140 or 141 or 161):
            name[ij] = 'i'
        elif ord(name[ij]) == (147 or 148 or 149 or 162):
            name[ij] = 'o'
        elif ord(name[ij]) == (129 or 150 or 151 or 163):
            name[ij] = 'u'
    name1 = name.split(' ')
    name1.append(name1[0])
    name1.__delitem__(0)
    name2 = name1.copy()
    name2.append(name2[0])
    name2.__delitem__(0)
    name1 = " ".join(name1)
    name2 = " ".join(name2)
    # print(name)
    # print(name1)
    # print(name2)

    lowball = 9
    if where == players:
        lowball = 0
    if size == 1:
        lowball = 0

    for i in range(lowball, size):
        if where == players:
            if where[i] != 0:
                temp17 = str(where[i][1])
            else:
                continue
        elif size == 1:
            temp17 = where
        else:
            temp17 = str(where["B" + str(i)].value)
        if not len(temp17) > 0:
            return 0
        temp18 = list(temp17.lower()).copy()
        # print(temp18)
        for ij in range(0, len(temp18)):
            if ord(temp18[ij]) == 135:
                temp18[ij] = 'c'
            elif ord(temp18[ij]) == (131 or 132 or 133 or 134 or 160):
                temp18[ij] = 'a'
            elif ord(temp18[ij]) == (130 or 136 or 137):
                temp18[ij] = 'e'
            elif ord(temp18[ij]) == (139 or 140 or 141 or 161):
                temp18[ij] = 'i'
            elif ord(temp18[ij]) == (147 or 148 or 149 or 162):
                temp18[ij] = 'o'
            elif ord(temp18[ij]) == (129 or 150 or 151 or 163):
                temp18[ij] = 'u'
        temp17 = "".join(temp18)
        # print(name, name1, name2, i, temp17, [rtgfnx["A" + str(i)].value, rtgfnx["B" + str(i)].value, rtgfnx["J" + str(i)].value])
        if countingwords(name, temp17) or countingwords(name1, temp17) or countingwords(name2, temp17):
            # print('find', [rtgfnx["A" + str(i)].value, temp17, rtgfnx["J" + str(i)].value])
            if size == 1:
                return True
            if hint is not None:
                return i
            if type(where) == list:
                return [where[i][0], where[i][1], where[i][2]]
            if where == players:
                return where[i]
            return [where["A" + str(i)].value, where["B" + str(i)].value, where["J" + str(i)].value]
        if (name1 in temp17) or (name2 in temp17) or (name1 in temp17):
            # print('find', [rtgfnx["A" + str(i)].value, temp17, rtgfnx["J" + str(i)].value])
            if size == 1:
                return True
            if hint is not None:
                return i
            # print("Debug:", ["A" + str(i), "B" + str(i), "J" + str(i)], where)
            if type(where) == list:
                return [where[i][0], where[i][1], where[i][2]]
            return [where["A" + str(i)].value, where["B" + str(i)].value, where["J" + str(i)].value]
    return 0


def findplayerinlist(name, player):
    """ Return the rating of given player in the tournament list, just don't forget to
    properly add a player, or it'll be defined as a new FNX chess player"""
    """ Return zero if the player does not have Rating in the player list"""
    for i in range(0, len(player)):
        if player[i] != 0:
            if player[i][1] == name:
                # print(name, player[i][2])
                return player[i][2]
    return 0


def ratingperformance(result, averagertg, perftable):
    """ Return Rating Performance based in  https://en.wikipedia.org/wiki/Performance_rating_(chess)"""
    if result == 0:
        return round(averagertg - 500, 2)
    if result == 9:
        return round(averagertg + 500, 2)
    if result == 4.5:
        return round(averagertg, 2)
    if result < 4.5:
        return round(averagertg + perftable[1][int(result / 0.5) - 1], 2)
    if result > 4.5:
        return round(averagertg + perftable[3][(int(result / 0.5)) - 9], 2)
    return 0


def round_off_rating(number):
    """ Reduces around 0.5 float numbers """
    return round(number * 2) / 2


""" The rating performance matrix, defined by FIDE in 
https://en.wikipedia.org/wiki/Performance_rating_(chess)"""
pm = np.zeros((4, 8), dtype=float)
np.set_printoptions(suppress=True)
pm[0][0] = 0.06
pm[0][1] = 0.11
pm[0][2] = 0.17
pm[0][3] = 0.22
pm[0][4] = 0.38
pm[0][5] = 0.33
pm[0][6] = 0.39
pm[0][7] = 0.44
pm[1][0] = -444
pm[1][1] = -351
pm[1][2] = -273
pm[1][3] = -220
pm[1][4] = -166
pm[1][5] = -125
pm[1][6] = -80
pm[1][7] = -43
pm[2][0] = 0.56
pm[2][1] = 0.61
pm[2][2] = 0.67
pm[2][3] = 0.72
pm[2][4] = 0.78
pm[2][5] = 0.83
pm[2][6] = 0.89
pm[2][7] = 0.94
pm[3][0] = 43
pm[3][1] = 80
pm[3][2] = 125
pm[3][3] = 166
pm[3][4] = 220
pm[3][5] = 273
pm[3][6] = 351
pm[3][7] = 444
# print(pm)


print("Starting")
# flipexcel(1)
tournamentstest = os.listdir('tournaments')
# print(tournamentstest)
for i in range(len(tournamentstest)):
    """ It's almost impossible to convert .xls to .xlsx"""
    # Microsoft is making my life harder
    time.sleep(1)
    if ".xlsx" in tournamentstest[i] and "Table_Starting" in tournamentstest[i]:
        print('\n', tournamentstest[i], 'NonValidFileType')
    elif ".xlsx" in tournamentstest[i] and "_FIDE_" in tournamentstest[i]:
        """ Find out the tournament type, if not found set to k = 10 (Please properly rename the files 
        before uploading"""
        if "std" in tournamentstest[i] or "standard" in tournamentstest[i] or "classical" in tournamentstest[i] or "classico" in tournamentstest[i] or "clássico" in tournamentstest[i] or "pensado" in tournamentstest[i]:
            kfactor = 0
        elif "rpd" in tournamentstest[i] or "rapid" in tournamentstest[i] or "rapido" in tournamentstest[i] in tournamentstest[i] or "rápido" in tournamentstest[i]:
            kfactor = 1
        else:  # If somewhat the tournament time control is not defined, set as blitz.
            kfactor = 2
        print('\n', tournamentstest[i], kfactor)
        # print("true")
        # Change .xls to .xlsx using https://convertio.co/pt/xls-xlsx/ or something

        """ Defining the path of the rtgfnx file which will read and write in."""
        xlsx_file1 = Path('fnxlist', 'rtgfnx.xlsx')
        wb_obj1 = openpyxl.load_workbook(xlsx_file1, data_only=True)
        rtgfnx = wb_obj1.active
        row_countfnx = rtgfnx.max_row
        column_countfnx = rtgfnx.max_column
        maxfnx = 0
        temp11 = 0

        """ Just to be sure that the maxfnx ID is updated after every change"""
        for la in range(9, row_countfnx):
            temp16 = rtgfnx["A" + str(la)].value
            if temp16 is not None:
                # print(temp16, type(temp16))
                if maxfnx < temp16:
                    maxfnx = int(temp16)
        # print('maxfnx:', maxfnx, row_countfnx)

        """ Defining the path of the tournament file that we r working with"""
        xlsx_file2 = Path('tournaments', tournamentstest[i])
        wb_obj2 = openpyxl.load_workbook(xlsx_file2, data_only=True)
        tournament = wb_obj2.active
        row_counttournament = tournament.max_row
        column_counttournament = tournament.max_column

        """ Making a backup of rtgfnx"""
        shutil.copy(Path('fnxlist', 'rtgfnx.xlsx'), 'backup')
        old_name = Path('backup', 'rtgfnx.xlsx')
        new_name = Path('backup', 'rtgfnx' + str(datetime.now().strftime("%Y_%m_%d_%I_%M_%S_%p")) + '.xlsx')
        os.rename(old_name, new_name)
        players = []

        """ Defining the path of the performance list, actually just a background file"""
        """ DO NOT EDIT THIS FILE MANUALLY"""
        xlsx_file3 = Path('fnxlist', 'perfrtgfnx.xlsx')
        wb_obj3 = openpyxl.load_workbook(xlsx_file3, data_only=True)
        perfrtgfnx = wb_obj3.active
        row_countperfrtgfnx = perfrtgfnx.max_row
        column_countperfrtgfnx = perfrtgfnx.max_column

        """ Making sure that all the files loaded properly"""
        # Yes I don't want to use subprocess.call
        time.sleep(3)

        for j in range(1, row_counttournament):
            """ Reading the tournament file"""
            # time.sleep(0.5)
            # print(j, row_countfnx)
            if tournament["E" + str(j)].value == "Name:":
                # print('\n')
                temp12 = findplayer(str(tournament["G" + str(j)].value), row_countfnx, rtgfnx)
                # print(temp12, tournament["G" + str(j)].value)
                """ Verify if the player subscribed in the tournament has FNX Rtg"""
                if temp12 != 0:
                    # print("True", temp12)
                    players.append(temp12)
                    # rating1 = temp12[2]
                else:
                    # temp20 = findplayerinlist(tournament["G" + str(j)].value, players)
                    # print("False", temp20)
                    players.append(0)
                    # rating1 = temp20
                # print('Rating1:', rating1)
                """ If the player does not have FNX rating and the player rating 
                was set manually(Please dont do this), 
                we save in the players list one of them"""
        # print(players)
        for j in range(1, row_counttournament):
            """ Reading the tournament file"""
            # time.sleep(0.5)
            # print(j, row_countfnx)
            if tournament["E" + str(j)].value == "Name:":
                # print('\n')
                temp12 = findplayer(str(tournament["G" + str(j)].value), row_countfnx, rtgfnx)
                # print(temp12, tournament["G" + str(j)].value)
                """ Verify if the player subscribed in the tournament has FNX Rtg"""
                # print(row_countfnx)
                if temp12 != 0:
                    # print("True", temp12)
                    # players.append(temp12)
                    rating1 = temp12[2]
                else:
                    # print("False", temp20)
                    # players.append(0)
                    rating1 = 0
                # print("Debug:", temp12, rating1)
                for l in range(1, 1000):
                    if tournament["A" + str(j + l)].value == "Rd.":
                        """ Reading the matchs results of the player"""
                        # print('\n')
                        variation = 0
                        """ Saving the 'Res.' cell coord """
                        if str(tournament["E" + str(j + l)].value) == 'Res.':
                            point = 'E'
                        elif str(tournament["F" + str(j + l)].value) == 'Res.':
                            point = 'F'
                        elif str(tournament["G" + str(j + l)].value) == 'Res.':
                            point = 'G'
                        elif str(tournament["H" + str(j + l)].value) == 'Res.':
                            point = 'H'
                        elif str(tournament["I" + str(j + l)].value) == 'Res.':
                            point = 'I'
                        elif str(tournament["J" + str(j + l)].value) == 'Res.':
                            point = 'J'
                        for m in range(1, 20):
                            # actually the limit is 20 because I don't believe will happen a tournament with this amount of games
                            if tournament["A" + str(j + l + m)].value is None:
                                # but this line is our safeproof breakrule
                                break
                            # print(players)
                            temp21 = findplayer(tournament["D" + str(j + l + m)].value, len(players), players)
                            if temp21 != 0:
                                rating2 = temp21[2]
                            else:
                                rating2 = 0
                            # print(tournament["D" + str(j + l + m)].value, rating2, j + l + m)
                            if rating1 == 0:
                                """ Which means player does not have FNX Rtg and need to be registed"""
                                perfc = 0
                                score = 0
                                n1 = 0
                                for n in range(1, 100):
                                    # print(j + l + n)
                                    if tournament["A" + str(j + l + n)].value is None:
                                        break
                                    temp24 = 0
                                    temp2 = findplayer(tournament["D" + str(j + l + n)].value, len(players), players)
                                    if temp2 != 0:
                                        temp24 = temp2[2]
                                    else:
                                        temp24 = 0
                                    # print(temp2, rating2)
                                    # temp2 = findplayerinlist(tournament["D" + str(j + l + n)].value, players)
                                    temp1 = str(tournament[str(point) + str(j + l + n)].value)
                                    # print(temp1, j + l, point)
                                    if ord(temp1) == 189:  # Actually ½ isn't a number for some reason
                                        temp1 = 0.5
                                    if temp1 != '0' and temp1 != 0.5 and temp1 != '1':
                                        temp1 = 0
                                    if temp24 != 0:
                                        n1 = n1 + 1
                                        perfc = perfc + temp24
                                        score = round(score + float(temp1), 1)
                                # print(perfc, score, n1)

                                # print(ratingperformance(score, perfc, pm))

                                # print('size', row_countperfrtgfnx, column_countperfrtgfnx)
                                test = False  # Check if player is in performance rating list
                                for o in range(1, row_countperfrtgfnx+1):
                                    # print("Row Count Fnx Debug: ", row_countfnx)
                                    # print('debug', tournament["G" + str(j)].value, perfrtgfnx["B" + str(o)].value)
                                    temp22 = str(perfrtgfnx["B" + str(o)].value)
                                    temp23 = findplayer(tournament["G" + str(j)].value, 1, temp22)
                                    # print(temp23, tournament["G" + str(j)].value, temp22)
                                    # print(temp23)
                                    # if tournament["G" + str(j)].value == str(perfrtgfnx["B" + str(o)].value):
                                    if temp23 != 0:
                                        """ Check if the player is in performance rating list"""
                                        # print(o)
                                        temp5 = round(float(perfrtgfnx["C" + str(o)].value) + n1, 2)
                                        # print("debug:", perfrtgfnx["J" + str(o)].value, score, temp5, o)
                                        if temp5 >= 9 or perfrtgfnx["C" + str(o)].value == -1:
                                            """ If the number of rated games, readed in performance rating list, 
                                            reaches 9 we evaluate his initial rating, erase his data from performance 
                                            rating list and add in rating fnx list"""
                                            """ I'm just erasing the cell and not compacting the remaining data, 
                                            one day i solve that"""
                                            # print("Debug:", temp5)
                                            # wb2 = openpyxl.load_workbook(xlsx_file1)
                                            # ws2 = wb2.active  # or wb.active
                                            temp15 = tournament["G" + str(j)].value
                                            d = False
                                            # print("Max Games1", temp15)
                                            for ia in range(9, row_countfnx + 1):
                                                if temp15 in str(rtgfnx["B" + str(ia)].value):
                                                    """ Just to be sure if someone didn't add manually"""
                                                    # print("debug", ia, row_countfnx, rtgfnx["B" + str(ia)].value, temp15)
                                                    d = True
                                            if not d:
                                                # print("Max Games2", temp15)
                                                rtgfnx['A' + str(row_countfnx + 1)] = int(
                                                    perfrtgfnx["A" + str(o)].value)
                                                rtgfnx['B' + str(row_countfnx + 1)] = str(
                                                    perfrtgfnx["B" + str(o)].value)
                                                if perfrtgfnx["C" + str(o)].value == -1:
                                                    rtgfnx['J' + str(row_countfnx + 1)] = int(
                                                        perfrtgfnx["D" + str(o)].value)
                                                else:
                                                    rtgfnx['J' + str(row_countfnx + 1)] = int(ratingperformance(float(
                                                        6 * int(perfrtgfnx["J" + str(o)].value) / (
                                                            int(perfrtgfnx["C" + str(o)].value))), round(
                                                        int(perfrtgfnx["D" + str(o)].value) / (
                                                            int(perfrtgfnx["C" + str(o)].value)), 0), pm))
                                                rtgfnx['CS' + str(row_countfnx + 1)] = 1
                                                row_countfnx = row_countfnx + 1
                                                perfrtgfnx['A' + str(o)] = None
                                                perfrtgfnx['B' + str(o)] = None
                                                perfrtgfnx['C' + str(o)] = None
                                                perfrtgfnx['D' + str(o)] = None
                                                perfrtgfnx['E' + str(o)] = None
                                                perfrtgfnx['J' + str(o)] = None
                                                # wb2.save(xlsx_file1)
                                        else:
                                            """ And if he didn't reached 9 games, just add the data to 
                                            performance rating list"""
                                            perfrtgfnx['C' + str(o)] = perfrtgfnx["C" + str(o)].value + n1
                                            perfrtgfnx['D' + str(o)] = perfrtgfnx["D" + str(o)].value + perfc
                                            perfrtgfnx['J' + str(o)] = perfrtgfnx["J" + str(o)].value + score
                                        test = True
                                        break
                                if not test:
                                    """ If we didn't find the player in the performance rating list"""
                                    """ So we add him"""
                                    # print('perfdata')
                                    # print(maxfnx)
                                    # worksheet.write('A3', 'teste')
                                    # workbook.close()
                                    # worksheet.write('D' + str(row_countperfrtgfnx), 'test1')
                                    # print('size', row_countperfrtgfnx, column_countperfrtgfnx)
                                    temp26 = int(tournament[
                                                     "G" + str(j + 1)].value)
                                    if temp26 is None:
                                        temp26 = 0
                                    test2 = 0
                                    if tournament["E" + str(j + 1)].value == "RatNat:" and temp26 > 0:
                                        test2 = 1
                                    perfrtgfnx['A' + str(row_countperfrtgfnx + 1)] = int(maxfnx + 1)
                                    maxfnx = maxfnx + 1
                                    perfrtgfnx['B' + str(row_countperfrtgfnx + 1)] = tournament["G" + str(j)].value
                                    if test2 == 1:
                                        perfrtgfnx['C' + str(row_countperfrtgfnx + 1)] = -1
                                        perfrtgfnx['D' + str(row_countperfrtgfnx + 1)] = temp26
                                        perfrtgfnx['J' + str(row_countperfrtgfnx + 1)] = -1
                                        perfrtgfnx['E' + str(row_countperfrtgfnx + 1)] = -1
                                    else:
                                        # temp4 = perfrtgfnx['C' + str(row_countperfrtgfnx + 1)].value
                                        # if type(temp4) == int:
                                        #     perfrtgfnx('C' + str(row_countperfrtgfnx + 1), n1 + int(temp4))
                                        # else:
                                        perfrtgfnx['C' + str(row_countperfrtgfnx + 1)] = n1
                                        perfrtgfnx['D' + str(row_countperfrtgfnx + 1)] = perfc
                                        perfrtgfnx['J' + str(row_countperfrtgfnx + 1)] = score
                                    row_countperfrtgfnx = row_countperfrtgfnx + 1
                                # wb2.save(xlsx_file1)
                                # wb.save(xlsx_file3)
                                break
                            else:
                                """ The player has FNX RTG and we just evaluate his variation"""
                                # print("debug:", m, rating1, rating2)
                                # print(rating1, rating2, j, l, m, j+l+m)
                                variation = round(
                                    variation + var(rating1, rating2, tournament[point + str(j + l + m)].value, kfactor)
                                    , 2)
                                # wb2 = openpyxl.load_workbook(xlsx_file1)
                                # ws2 = wb2.active  # or wb.active

                            # wb2.save(xlsx_file1)
                        temp14 = tournament["G" + str(j)].value
                        # print(temp14, variation)
                        if rating1 != 0:
                            """ If the player has rating, obviusly is in FNX RTG and we just update his data"""
                            d = findplayer(temp14, row_countfnx, rtgfnx, 'supersecretparameter')
                            print(d, temp14, rtgfnx["J" + str(d)].value, variation)
                            rtgfnx['J' + str(d)] = round(int(rtgfnx["J" + str(d)].value) + variation, 0)
                            temp19 = rtgfnx["CS" + str(d)].value
                            if temp19 is None:
                                temp19 = 0
                            rtgfnx['CS' + str(d)] = int(temp19) + 1
                        break
        """ Saving the files before opening the next tournament file"""
        wb_obj1.save(xlsx_file1)
        wb_obj3.save(xlsx_file3)
        # wb2.save(xlsx_file1)
        # print(players)
        """ Move the tournament file to the 'already readed' folder """
        shutil.move(xlsx_file2, Path('savedtournaments', tournamentstest[i]))
        # quit()
        """ Some assurements about the tournament: """
        """ Some issues can be found due to players names beeing a substring of others names, indeed check manually "
              "if the user FNX ID is updated as expected after the tournament, if it doesnt happen ask @Davi Nascimento for "
              "and solution or set manually the players parameters somewhere around line 358 near major j loop, temp12 and"
              "tournament cell equals 'Name:'. """
    else:
        print('\n', tournamentstest[i], "Isn't a tournament file")

duration = 1000  # milliseconds
freq = 440  # Hz
winsound.Beep(freq, duration)
print("end")
